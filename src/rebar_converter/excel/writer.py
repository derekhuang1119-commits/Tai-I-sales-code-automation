from pathlib import Path

from rebar_converter.models import RebarItem


HEADERS = ("區域", "頁數", "鋼種", "號數", "左上", "左長", "左下", "中上",
           "中下", "右上", "右長", "右下", "鳥嘴", "總長", "支數", "總重")
FIELDS = ("region", "page", "steel_grade", "bar_number", "left_top", "left_long",
          "left_bottom", "middle_top", "middle_bottom", "right_top", "right_long",
          "right_bottom", "has_bird_beak", "total_length", "quantity", "total_weight")
# Excel's localized representation for a true bird beak flag.
BIRD_BEAK_MARK = "是"
HEADER_ROW = 1


class ExcelWriter:
    def write(self, items: list[RebarItem], output_path: Path, template_path: Path | None = None) -> Path:
        if template_path and output_path.resolve() == template_path.resolve():
            raise ValueError("輸出檔案不可覆蓋 Excel 範本")
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import PatternFill
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for Excel output") from exc
        if template_path:
            workbook = load_workbook(template_path)
        else:
            workbook = Workbook()
        sheet = workbook.active
        headers = [cell.value for cell in sheet[HEADER_ROW]]
        if not any(header in HEADERS for header in headers):
            for column, header in enumerate(HEADERS, 1):
                sheet.cell(HEADER_ROW, column, header)
            headers = list(HEADERS)
        else:
            for header in HEADERS:
                if header not in headers:
                    headers.append(header)
                    sheet.cell(HEADER_ROW, len(headers), header)
        positions = {header: index + 1 for index, header in enumerate(headers) if header in HEADERS}
        review_fill = PatternFill("solid", fgColor="FFF2CC")
        for row, item in enumerate(items, 2):
            for header, field in zip(HEADERS, FIELDS):
                column = positions.get(header)
                if column:
                    value = getattr(item, field)
                    sheet.cell(row, column, BIRD_BEAK_MARK if value is True else value)
                    if item.needs_review:
                        sheet.cell(row, column).fill = review_fill
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path
