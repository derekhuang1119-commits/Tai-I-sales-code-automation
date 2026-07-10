from pathlib import Path

import pytest
from openpyxl import load_workbook

from tai_i_sales.excel import write_excel
from tai_i_sales.ingest import image_to_page, pdf_to_pages
from tai_i_sales.models import BarItem, OCRToken, PageOCR
from tai_i_sales.ocr import OCRModelUnavailable
from tai_i_sales.parsing import page_marker_number, parse_row, row_rois
from tai_i_sales.shape import analyze_shape, row_is_crossed_out


def token(text, x, y, confidence=0.95, page=1):
    return OCRToken(text, confidence, ((x, y), (x + 20, y), (x + 20, y + 10), (x, y + 10)), page)


class FakeOCR:
    def __init__(self):
        self.calls = []

    def recognize(self, image, page_number):
        self.calls.append((image, page_number))
        return [token("3-1", 900, 900, page=page_number)]


def test_image_uses_injected_ocr_backend():
    from PIL import Image

    backend = FakeOCR()
    page = image_to_page(Image.new("RGB", (1000, 1000)), backend)
    assert backend.calls and page.tokens[0].polygon


def test_scanned_pdf_renders_300_dpi_and_uses_ocr(monkeypatch):
    import fitz

    class Pix:
        def tobytes(self, _):
            from PIL import Image
            import io
            output = io.BytesIO()
            Image.new("RGB", (100, 100)).save(output, "PNG")
            return output.getvalue()

    class PdfPage:
        rect = fitz.Rect(0, 0, 72, 72)

        def get_text(self, _):
            return ""

        def get_pixmap(self, matrix, alpha):
            assert matrix.a == pytest.approx(300 / 72)
            return Pix()

    class Document:
        def __iter__(self):
            return iter([PdfPage()])

    backend = FakeOCR()
    pages = pdf_to_pages("synthetic.pdf", backend, pdf_open=lambda _: Document(), text_threshold=20)
    assert pages[0].used_ocr_fallback
    assert len(backend.calls) == 1


def test_page_marker_uses_bottom_right_dash_value():
    page = PageOCR(1, 1000, 1000, [token("3-1", 850, 850), token("3-2", 850, 900)])
    assert page_marker_number(page) == 1  # first matching marker is the page marker


def test_x_in_ocr_text_does_not_exclude_row():
    item = parse_row(
        [token("X", 10, 10), token("號數:#12", 30, 10), token("總長:600", 50, 10),
         token("支數:2", 70, 10), token("總重:30", 90, 10)],
        region="北區", page_number=1,
    )
    assert not item.excluded
    assert item.bar_number == "12"


def test_rows_are_split_by_horizontal_table_lines_and_retain_geometry():
    page = PageOCR(1, 200, 200, [token("a", 10, 20), token("b", 10, 80)])
    rows = row_rois(page, [0, 50, 100])
    assert [row[0].text for row in rows] == ["a", "b"]
    assert rows[0][0].bbox[0] == 10


def test_straight_bar_goes_to_middle_top_and_non_stirrup_bird_mouth_empty():
    shape = analyze_shape([token("120", 50, 50)], is_stirrup=False)
    assert shape.middle_top == "120"
    assert shape.bird_mouth == ""


def test_normal_text_x_is_not_a_crossout_signal():
    assert row_is_crossed_out(None, (0, 0, 10, 10)) is None


def test_excel_maps_by_titles_and_leaves_steel_blank(tmp_path: Path):
    workbook = __import__("openpyxl").Workbook()
    sheet = workbook.active
    for col, title in enumerate(["總重", "區域", "鋼種", "頁數", "號數", "支數", "總長"], 1):
        sheet.cell(1, col, title)
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    workbook.save(template)
    write_excel([BarItem(region="北區", page_number=1, bar_number="12", total_length="600",
                         quantity="2", total_weight="30")], output, template)
    result = load_workbook(output).active
    assert result.cell(2, 2).value == "北區"
    assert result.cell(2, 3).value is None
    assert result.cell(2, 7).value == "600"


def test_excel_rejects_missing_required_fields(tmp_path):
    from openpyxl import Workbook
    template = tmp_path / "template.xlsx"
    sheet = Workbook()
    sheet.active.append(["區域", "頁數", "號數", "總長", "支數", "總重", "鋼種"])
    sheet.save(template)
    with pytest.raises(ValueError, match="必要"):
        write_excel([BarItem(region="北區")], tmp_path / "output.xlsx", template)
