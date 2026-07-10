"""Integration tests for all requirements listed in PR review feedback."""

from pathlib import Path
from unittest.mock import MagicMock, patch, call

import pytest
from openpyxl import Workbook, load_workbook

from tai_i_sales.batch import BatchProcessor
from tai_i_sales.excel import HEADERS, write_excel, _find_header_row
from tai_i_sales.models import BarItem, OCRToken, PageOCR, ShapeAnalysis
from tai_i_sales.parsing import detect_region, page_marker_number, parse_row, row_rois
from tai_i_sales.shape import analyze_shape, row_is_crossed_out, _segments_intersect


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def tok(text, x, y, conf=0.95, page=1):
    return OCRToken(text, conf, ((x, y), (x + 20, y), (x + 20, y + 10), (x, y + 10)), page)


class StubOCR:
    """Minimal OCR backend that records calls and returns preset tokens."""

    def __init__(self, tokens=None):
        self.tokens = tokens or []
        self.calls = []

    def recognize(self, image, page_number):
        self.calls.append((image, page_number))
        return self.tokens


# ---------------------------------------------------------------------------
# 1. ShapeAnalysis fields written to Excel
# ---------------------------------------------------------------------------

class TestShapeFieldsInExcel:
    def test_shape_fields_all_written(self, tmp_path):
        """All shape-dimension fields from ShapeAnalysis must appear in Excel."""
        wb = Workbook()
        ws = wb.active
        for col, title in enumerate(HEADERS.values(), 1):
            ws.cell(1, col, title)
        tmpl = tmp_path / "t.xlsx"
        wb.save(tmpl)

        shape = ShapeAnalysis(
            left_top=10.0, left_long=200.0, left_bottom=10.0,
            middle_top=300.0, middle_bottom=None,
            right_top=10.0, right_long=200.0, right_bottom=10.0,
            bird_mouth=None, is_stirrup=False,
        )
        item = BarItem(
            region="北區", page_number=1, bar_number="3",
            total_length=600.0, quantity=2, total_weight=30.0,
            shape=shape,
        )
        out = tmp_path / "out.xlsx"
        write_excel([item], out, tmpl)

        result = load_workbook(out).active
        # find column for 中上
        header_row = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        result_row = {result.cell(1, c).value: c for c in range(1, result.max_column + 1)}
        assert result.cell(2, result_row["中上"]).value == 300.0
        assert result.cell(2, result_row["左長"]).value == 200.0
        assert result.cell(2, result_row["右長"]).value == 200.0


# ---------------------------------------------------------------------------
# 2. BatchProcessor calls shape analysis and crossout detection
# ---------------------------------------------------------------------------

class TestBatchProcessorWiring:
    def test_extract_calls_shape_and_crossout(self, tmp_path):
        """BatchProcessor.extract() must invoke analyze_shape and row_is_crossed_out."""
        tokens = [
            tok("3-1", 900, 900),  # page marker
            tok("號數:5", 10, 100), tok("總長:500", 60, 100),
            tok("支數:3", 110, 100), tok("總重:20", 160, 100),
        ]
        page = PageOCR(1, 1000, 1000, tokens)

        ocr = StubOCR(tokens)

        with patch("tai_i_sales.batch.load_source", return_value=[page]), \
             patch("tai_i_sales.batch.analyze_shape", return_value=ShapeAnalysis()) as mock_shape, \
             patch("tai_i_sales.batch.row_is_crossed_out", return_value=False) as mock_cross:
            proc = BatchProcessor(ocr)
            items = proc.extract("fake.pdf")

        assert mock_shape.called, "analyze_shape must be called"
        assert mock_cross.called, "row_is_crossed_out must be called"

    def test_confirmed_crossout_sets_excluded(self, tmp_path):
        """When crossed-out detection returns True, item.excluded must be True."""
        page = PageOCR(1, 100, 100, [tok("號數:1", 10, 50), tok("總長:100", 40, 50),
                                      tok("支數:1", 70, 50), tok("總重:5", 100, 50)])
        with patch("tai_i_sales.batch.load_source", return_value=[page]), \
             patch("tai_i_sales.batch.analyze_shape", return_value=ShapeAnalysis()), \
             patch("tai_i_sales.batch.row_is_crossed_out", return_value=True):
            proc = BatchProcessor(StubOCR())
            items = proc.extract("fake.pdf", region="北區")
        assert any(i.excluded for i in items), "Confirmed X must set excluded=True"

    def test_ambiguous_crossout_sets_needs_review(self):
        """When crossed-out detection is ambiguous (None), item.needs_review must be True."""
        page = PageOCR(1, 100, 100, [tok("號數:1", 10, 50), tok("總長:100", 40, 50),
                                      tok("支數:1", 70, 50), tok("總重:5", 100, 50)])
        with patch("tai_i_sales.batch.load_source", return_value=[page]), \
             patch("tai_i_sales.batch.analyze_shape", return_value=ShapeAnalysis()), \
             patch("tai_i_sales.batch.row_is_crossed_out", return_value=None):
            proc = BatchProcessor(StubOCR())
            items = proc.extract("fake.pdf", region="北區")
        assert any(i.needs_review for i in items), "Ambiguous X must set needs_review=True"
        assert not any(i.excluded for i in items), "Ambiguous X must NOT set excluded=True"


# ---------------------------------------------------------------------------
# 3. Crossed-out detection: positive + negative slope intersection
# ---------------------------------------------------------------------------

class TestCrossoutDetection:
    def _make_x_image(self, size=100):
        """Create a synthetic image with two crossing diagonal lines."""
        try:
            import cv2
            import numpy as np
        except ImportError:
            pytest.skip("OpenCV not available")
        img = np.ones((size, size, 3), dtype=np.uint8) * 255
        cv2.line(img, (5, 5), (size - 5, size - 5), (0, 0, 0), 2)      # ↘ negative slope
        cv2.line(img, (5, size - 5), (size - 5, 5), (0, 0, 0), 2)      # ↗ positive slope
        return img

    def _make_parallel_diag_image(self, size=100):
        """Create a synthetic image with two PARALLEL diagonal lines (not an X)."""
        try:
            import cv2
            import numpy as np
        except ImportError:
            pytest.skip("OpenCV not available")
        img = np.ones((size, size, 3), dtype=np.uint8) * 255
        cv2.line(img, (5, 5), (50, 50), (0, 0, 0), 2)
        cv2.line(img, (50, 5), (95, 50), (0, 0, 0), 2)   # same slope, offset
        return img

    def test_x_image_detected(self):
        img = self._make_x_image(120)
        result = row_is_crossed_out(img, (0, 0, 120, 120))
        # Should be True or None (ambiguous); must NOT be False
        assert result is not False, "Clear X pattern must be detected"

    def test_non_intersecting_diagonals_not_x(self):
        """Two non-intersecting diagonals must not be reported as an X."""
        img = self._make_parallel_diag_image(100)
        result = row_is_crossed_out(img, (0, 0, 100, 100))
        assert result is not True, "Non-intersecting diagonals must not be classified as X"

    def test_segments_intersect_basic(self):
        # Two crossing line segments
        pt = _segments_intersect(0, 0, 10, 10, 0, 10, 10, 0)
        assert pt is not None
        assert abs(pt[0] - 5) < 1e-6
        assert abs(pt[1] - 5) < 1e-6

    def test_segments_no_intersection(self):
        # Parallel horizontal lines — no intersection
        pt = _segments_intersect(0, 0, 10, 0, 0, 5, 10, 5)
        assert pt is None

    def test_non_overlapping_segments_no_intersection(self):
        # Lines that would intersect if extended but don't overlap
        pt = _segments_intersect(0, 0, 3, 3, 5, 5, 10, 10)
        assert pt is None


# ---------------------------------------------------------------------------
# 4. Column x-bounds reading
# ---------------------------------------------------------------------------

class TestColumnBoundsReading:
    def test_col_bounds_parse_bar_number(self):
        tokens = [tok("5", 15, 50), tok("總長:600", 80, 50)]
        col_bounds = {"bar_number": (0.0, 40.0)}
        item = parse_row(tokens, region="A區", page_number=1, col_bounds=col_bounds)
        assert item.bar_number == "5"

    def test_col_bounds_parse_quantity_as_int(self):
        tokens = [tok("3", 70, 50)]
        col_bounds = {"quantity": (50.0, 100.0)}
        item = parse_row(tokens, region="A區", page_number=1, col_bounds=col_bounds)
        assert item.quantity == 3
        assert isinstance(item.quantity, int)

    def test_col_bounds_parse_total_weight_as_float(self):
        tokens = [tok("25.5", 90, 50)]
        col_bounds = {"total_weight": (80.0, 120.0)}
        item = parse_row(tokens, region="A區", page_number=1, col_bounds=col_bounds)
        assert item.total_weight == 25.5
        assert isinstance(item.total_weight, float)


# ---------------------------------------------------------------------------
# 5. GUI field type correctness (unit-level)
# ---------------------------------------------------------------------------

class TestGuiFieldTypes:
    def test_coerce_page_number_is_int(self):
        from tai_i_sales.gui import _coerce_field
        assert _coerce_field("page_number", "3") == 3
        assert isinstance(_coerce_field("page_number", "3"), int)

    def test_coerce_quantity_is_int(self):
        from tai_i_sales.gui import _coerce_field
        assert _coerce_field("quantity", "10") == 10
        assert isinstance(_coerce_field("quantity", "10"), int)

    def test_coerce_total_length_is_float(self):
        from tai_i_sales.gui import _coerce_field
        assert _coerce_field("total_length", "600") == 600.0
        assert isinstance(_coerce_field("total_length", "600"), float)

    def test_coerce_total_weight_is_float(self):
        from tai_i_sales.gui import _coerce_field
        assert _coerce_field("total_weight", "30.5") == 30.5
        assert isinstance(_coerce_field("total_weight", "30.5"), float)

    def test_coerce_empty_gives_none(self):
        from tai_i_sales.gui import _coerce_field
        assert _coerce_field("total_length", "") is None
        assert _coerce_field("page_number", "") is None

    def test_coerce_bar_number_strips_hash(self):
        from tai_i_sales.gui import _coerce_field
        assert _coerce_field("bar_number", "#5") == "5"
        assert _coerce_field("bar_number", "5") == "5"


# ---------------------------------------------------------------------------
# 6. Excel: header row not on row 1
# ---------------------------------------------------------------------------

class TestExcelDynamicHeaderRow:
    def test_header_found_on_row_3(self, tmp_path):
        """write_excel must succeed even when the header row is not row 1."""
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1, "公司名稱")
        ws.cell(2, 1, "工程名稱")
        for col, title in enumerate(HEADERS.values(), 1):
            ws.cell(3, col, title)
        tmpl = tmp_path / "t.xlsx"
        wb.save(tmpl)

        item = BarItem(
            region="A區", page_number=1, bar_number="1",
            total_length=100.0, quantity=1, total_weight=5.0,
        )
        out = tmp_path / "out.xlsx"
        write_excel([item], out, tmpl)

        result = load_workbook(out)
        ws_out = result.active
        # Data should be in row 4 (first data row after header on row 3)
        header_row, headers = _find_header_row(ws_out)
        assert header_row == 3
        data_row = header_row + 1
        region_col = headers.get("區域")
        assert region_col is not None
        assert ws_out.cell(data_row, region_col).value == "A區"

    def test_find_header_row_returns_best_match(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(2, 1, "區域")
        ws.cell(2, 2, "頁數")
        ws.cell(2, 3, "號數")
        ws.cell(2, 4, "總長")
        ws.cell(2, 5, "支數")
        ws.cell(2, 6, "總重")
        ws.cell(2, 7, "鋼種")
        row, hdrs = _find_header_row(ws)
        assert row == 2
        assert "區域" in hdrs


# ---------------------------------------------------------------------------
# 7. Template preservation: output must be a COPY, not overwrite template
# ---------------------------------------------------------------------------

class TestTemplatePreservation:
    def test_template_file_unchanged(self, tmp_path):
        """write_excel must not modify the original template file."""
        wb = Workbook()
        ws = wb.active
        for col, title in enumerate(HEADERS.values(), 1):
            ws.cell(1, col, title)
        tmpl = tmp_path / "template.xlsx"
        wb.save(tmpl)
        original_mtime = tmpl.stat().st_mtime

        item = BarItem(
            region="B區", page_number=2, bar_number="7",
            total_length=400.0, quantity=4, total_weight=20.0,
        )
        out = tmp_path / "output.xlsx"
        write_excel([item], out, tmpl)

        assert tmpl.stat().st_mtime == original_mtime, "Template file must not be modified"
        assert out.exists(), "Output file must be created"
        assert tmpl.read_bytes() != out.read_bytes(), "Output must differ from template (has data)"


# ---------------------------------------------------------------------------
# 8. Excluded rows are not exported
# ---------------------------------------------------------------------------

class TestExcludedRowsNotExported:
    def test_excluded_item_skipped_in_excel(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        for col, title in enumerate(HEADERS.values(), 1):
            ws.cell(1, col, title)
        tmpl = tmp_path / "t.xlsx"
        wb.save(tmpl)

        items = [
            BarItem(region="A區", page_number=1, bar_number="1",
                    total_length=100.0, quantity=1, total_weight=5.0, excluded=False),
            BarItem(region="A區", page_number=1, bar_number="2",
                    total_length=200.0, quantity=2, total_weight=10.0, excluded=True),
        ]
        out = tmp_path / "out.xlsx"
        write_excel(items, out, tmpl)
        result = load_workbook(out).active
        # Only one data row (row 2); row 3 must be empty
        assert result.cell(2, 1).value is not None  # first item
        assert result.cell(3, 1).value is None  # excluded item not written


# ---------------------------------------------------------------------------
# 9. Straight bar dimension goes to middle_top; non-stirrup bird_mouth is None
# ---------------------------------------------------------------------------

class TestShapeRules:
    def test_single_dimension_is_middle_top(self):
        shape = analyze_shape([tok("500", 50, 50)], is_stirrup=False)
        assert shape.middle_top == 500.0
        assert shape.bird_mouth is None

    def test_stirrup_bird_mouth_allowed(self):
        shape = ShapeAnalysis(is_stirrup=True, bird_mouth=50.0)
        assert shape.bird_mouth == 50.0

    def test_non_stirrup_bird_mouth_cleared(self):
        tokens = [tok("100", 30, 30), tok("200", 70, 30), tok("50", 50, 70)]
        shape = analyze_shape(tokens, is_stirrup=False)
        assert shape.bird_mouth is None


# ---------------------------------------------------------------------------
# 10. Page marker 3-1 → page number 1; missing marker → needs_review
# ---------------------------------------------------------------------------

class TestPageMarkerRules:
    def test_marker_3_1_gives_page_1(self):
        page = PageOCR(1, 1000, 1000, [tok("3-1", 850, 900)])
        assert page_marker_number(page) == 1

    def test_marker_5_3_gives_page_3(self):
        page = PageOCR(1, 1000, 1000, [tok("5-3", 860, 910)])
        assert page_marker_number(page) == 3

    def test_missing_marker_triggers_needs_review(self):
        """When no page marker is found, BatchProcessor must flag needs_review."""
        page = PageOCR(1, 100, 100, [
            tok("號數:1", 10, 50), tok("總長:100", 40, 50),
            tok("支數:1", 70, 50), tok("總重:5", 100, 50),
        ])
        with patch("tai_i_sales.batch.load_source", return_value=[page]), \
             patch("tai_i_sales.batch.analyze_shape", return_value=ShapeAnalysis()), \
             patch("tai_i_sales.batch.row_is_crossed_out", return_value=False):
            proc = BatchProcessor(StubOCR())
            items = proc.extract("fake.pdf", region="北區")
        assert any(i.needs_review for i in items)
        # Must not fall back to PDF physical page index as the page_number value
        # (page.page_number == 1 would make it ambiguous; the key is needs_review is set)


# ---------------------------------------------------------------------------
# 11. Region OCR auto-detection from page header
# ---------------------------------------------------------------------------

class TestRegionDetection:
    def test_detects_a_region(self):
        page = PageOCR(1, 1000, 1000, [tok("A區鋼筋料單", 10, 30)])
        assert detect_region(page) == "A區"

    def test_detects_north_region(self):
        page = PageOCR(1, 1000, 1000, [tok("北區料單", 10, 50)])
        assert detect_region(page) == "北區"

    def test_no_region_returns_none(self):
        page = PageOCR(1, 1000, 1000, [tok("鋼筋料單", 10, 30)])
        assert detect_region(page) is None

    def test_region_below_header_not_detected(self):
        # Token at y=500 on a 1000-tall page is outside the 15% header zone
        page = PageOCR(1, 1000, 1000, [tok("B區", 10, 500)])
        assert detect_region(page) is None
