"""Excel template mapping by header title rather than fixed column numbers."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from .models import REQUIRED_FIELDS, BarItem


HEADERS: dict[str, str] = {
    "region": "區域",
    "page_number": "頁數",
    "steel_grade": "鋼種",
    "bar_number": "號數",
    "left_top": "左上",
    "left_long": "左長",
    "left_bottom": "左下",
    "middle_top": "中上",
    "middle_bottom": "中下",
    "right_top": "右上",
    "right_long": "右長",
    "right_bottom": "右下",
    "bird_mouth": "鳥嘴",
    "total_length": "總長",
    "quantity": "支數",
    "total_weight": "總重",
}

# Shape dimension fields sourced from ShapeAnalysis
_SHAPE_FIELDS = frozenset({
    "left_top", "left_long", "left_bottom",
    "middle_top", "middle_bottom",
    "right_top", "right_long", "right_bottom",
    "bird_mouth",
})

# Required field labels for validation
REQUIRED_HEADER_TITLES = {HEADERS[k] for k in REQUIRED_FIELDS}


def _find_header_row(sheet: Any) -> tuple[int, dict[str, int]]:
    """Find the header row and return (row_index, {title: col_index}).

    Searches every row for a row that contains the majority of known header
    titles. Returns the first matching row and never assumes row 1.
    """
    all_titles = set(HEADERS.values())
    best_row = 1
    best_headers: dict[str, int] = {}
    for row_idx in range(1, sheet.max_row + 1):
        row_cells = sheet[row_idx]
        found: dict[str, int] = {
            cell.value: cell.column
            for cell in row_cells
            if cell.value in all_titles
        }
        if len(found) > len(best_headers):
            best_headers = found
            best_row = row_idx
            if len(found) >= len(all_titles):
                break
    return best_row, best_headers


def _find_first_empty_row(sheet: Any, header_row: int) -> int:
    """Return the first row after header_row where all cells are empty."""
    for row_idx in range(header_row + 1, sheet.max_row + 2):
        row_cells = sheet[row_idx]
        if all(cell.value is None for cell in row_cells):
            return row_idx
    return sheet.max_row + 1


def write_excel(
    items: list[BarItem],
    destination: str | Path,
    template: str | Path | None = None,
) -> Path:
    """Write rows beneath matching Chinese headers; search header row dynamically.

    When a *template* path is supplied the workbook is loaded from it and saved
    to *destination*, leaving the template file untouched. When no template is
    provided a minimal workbook with all required headers is created.

    The function searches every row of the active sheet to locate the header
    row instead of assuming it is always the first row.
    """

    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel 支援需要安裝 openpyxl。") from exc

    if template is not None:
        workbook = load_workbook(str(template))
    else:
        workbook = Workbook()

    sheet = workbook.active

    if template is None and sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
        for column, title in enumerate(HEADERS.values(), 1):
            sheet.cell(row=1, column=column, value=title)

    header_row, headers = _find_header_row(sheet)

    # Validate that all required headers are present
    missing_required = [title for title in sorted(REQUIRED_HEADER_TITLES) if title not in headers]
    if missing_required:
        raise ValueError(f"Excel 範本缺少必要欄位標題：{', '.join(missing_required)}")

    for item in items:
        if item.excluded:
            continue
        missing = [HEADERS[key] for key in REQUIRED_FIELDS if getattr(item, key) in (None, "")]
        if missing:
            raise ValueError(f"料件缺少必要欄位：{', '.join(missing)}")

        first_empty = _find_first_empty_row(sheet, header_row)
        for field_key, title in HEADERS.items():
            if title not in headers:
                continue
            if field_key in _SHAPE_FIELDS:
                value = getattr(item.shape, field_key, None)
            else:
                value = getattr(item, field_key, None)
            # Write None as blank; never coerce to string
            if value == "":
                value = None
            sheet.cell(row=first_empty, column=headers[title], value=value)

    path = Path(destination)
    workbook.save(path)
    return path
